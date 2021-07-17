# Importing the Modules
from openpyxl.styles import Alignment, Font
import openpyxl
import requests

infoType = input('How do you want the information? (excel or console?) ')

if infoType == "excel":

    # Asking the user for the file name
    fileName = input('What do you want the file to be called? ')
    if len(fileName) < 1:
        fileName = "CryptoPrices"

    # Getting the requests from the coingecko API
    btcRaw = requests.get("https://api.coingecko.com/api/v3/simple/price?ids=bitcoin&vs_currencies=USD").json()
    btcPrice = btcRaw["bitcoin"]["usd"]
    btc = str(btcPrice)
    print('\n\nRecieved BTC\'s Price.')

    ethRaw = requests.get("https://api.coingecko.com/api/v3/simple/price?ids=ethereum&vs_currencies=USD").json()
    ethPrice = ethRaw["ethereum"]["usd"]
    eth = str(ethPrice)
    print('Recieved ETH\'s Price.')

    dogeRaw = requests.get("https://api.coingecko.com/api/v3/simple/price?ids=dogecoin&vs_currencies=USD").json()
    dogePrice = dogeRaw["dogecoin"]["usd"]
    doge = str(dogePrice)
    print('Recieved DOGE\'s Price.')

    adaRaw = requests.get("https://api.coingecko.com/api/v3/simple/price?ids=cardano&vs_currencies=USD").json()
    adaPrice = adaRaw["cardano"]["usd"]
    ada = str(adaPrice)
    print('Recieved ADA\'s Price.')

    # Initializing the workbook
    print('Creating WorkBook...')
    wb = openpyxl.Workbook()
    dest_filename = f'{fileName}.xlsx'
    ws = wb.active

    # Changing values of cells
    ws["A2"] = "Price"
    ws["B1"] = "BTC"
    ws["C1"] = "ETH"
    ws["D1"] = "DOGE"
    ws["E1"] = "ADA"
    ws["B2"] = btcPrice
    ws["C2"] = ethPrice
    ws["D2"] = dogePrice
    ws["E2"] = adaPrice
    ws.merge_cells('A3:D3')
    ws["A3"] = "ALL PRICES ARE IN USD!"
    ws['A3'].font = Font(bold=True)
    ws["A3"].alignment = Alignment(horizontal='center')

    # Saving the File!
    wb.save(f'{fileName}.xlsx')
    print('\nDone!')

elif infoType == "console":

    # Getting the requests from the coingecko API
    btcRaw = requests.get("https://api.coingecko.com/api/v3/simple/price?ids=bitcoin&vs_currencies=USD").json()
    btcPrice = btcRaw["bitcoin"]["usd"]
    btc = str(btcPrice)
    print('\n\nRecieved BTC\'s Price.')

    ethRaw = requests.get("https://api.coingecko.com/api/v3/simple/price?ids=ethereum&vs_currencies=USD").json()
    ethPrice = ethRaw["ethereum"]["usd"]
    eth = str(ethPrice)
    print('Recieved ETH\'s Price.')

    dogeRaw = requests.get("https://api.coingecko.com/api/v3/simple/price?ids=dogecoin&vs_currencies=USD").json()
    dogePrice = dogeRaw["dogecoin"]["usd"]
    doge = str(dogePrice)
    print('Recieved DOGE\'s Price.')

    adaRaw = requests.get("https://api.coingecko.com/api/v3/simple/price?ids=cardano&vs_currencies=USD").json()
    adaPrice = adaRaw["cardano"]["usd"]
    ada = str(adaPrice)
    print('Recieved ADA\'s Price.')


    
    print(f"\n\n\nBitcoin: ${btc}")
    print(f"Ethereum: ${eth}")
    print(f"Dogecoin: ${doge}")
    print(f"Cardano: ${ada}")

else:
    print("\n\nPlease Specify a valid Information Type!")